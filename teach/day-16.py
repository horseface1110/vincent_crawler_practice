import openpyxl

workbook = openpyxl.load_workbook('OwO.xlsx')
print(workbook.sheetnames)

print(workbook['sheetname1'])
mxR = sheet.max_row
